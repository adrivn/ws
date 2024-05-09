import datetime
import glob
import json
import re
import time
import os
from pathlib import Path
from typing import List, Tuple

from conf.settings import FileSettings
import duckdb
import openpyxl
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from rich.console import Console

console = Console()


def load_json_config(file):
    """
    Loads a JSON file containing external config data, like cell addresses and their corresponding output labels.
    """
    with open(file, "r", encoding="utf8") as f:
        return json.load(f)


def find_files_included(
    directory: Path,
    include_pattern: str,
    database_file: str,
    only_new_files: bool = False,
) -> Tuple[List[str], int]:
    query = f"""select full_path from ws.ws_offers_{include_pattern};"""

    all_files = []
    for subdir in directory.iterdir():
        if subdir.is_dir() and re.search(include_pattern, str(subdir)):
            console.print(f"Scanning offers directory {subdir.as_posix()}...")
            files = glob.iglob(subdir.as_posix() + "/**/[!~$]*.xlsx", recursive=True)
            all_files.extend(files)

    if only_new_files:
        with duckdb.connect(database_file) as db:
            ficheros_ya_leidos = [
                os.path.normpath(x[0]) for x in db.sql(query).fetchall()
            ]
            ficheros_existentes = [os.path.normpath(n) for n in all_files]
            ficheros_sin_leer = list(set(ficheros_existentes) - set(ficheros_ya_leidos))
            return ficheros_sin_leer, len(all_files)

    return [os.path.normpath(f) for f in all_files], len(all_files)


def auto_format_cell_width(ws):
    for letter in range(1, ws.max_column):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                maximum_value = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2


def create_style(json_file) -> dict[str, None]:
    # Load styles from JSON file
    with open(json_file) as file:
        style_data = json.load(file)

    style_dict = {}
    for k, v in style_data.items():
        if "font" in v and "fill" in v and "alignment" in v and "number_format" in v:
            # Create Named Style
            named_style = NamedStyle(name=k)

            # Set font
            named_style.font = Font(
                name=v["font"]["name"],
                size=v["font"]["size"],
                bold=v["font"]["bold"],
                italic=v["font"]["italic"],
                color=v["font"]["color"],
            )

            # Set fill
            named_style.fill = PatternFill(
                patternType=v["fill"]["pattern"], fgColor=v["fill"]["fgColor"]
            )

            # Set alignment
            named_style.alignment = Alignment(
                horizontal=v["alignment"]["horizontal"],
                vertical=v["alignment"]["vertical"],
            )

            # Set number format
            named_style.number_format = v["number_format"]

            # Set wrap_text if it is defined
            if "wrap_text" in v:
                named_style.alignment.wrap_text = v["wrap_text"]

            style_dict[k] = named_style

    return style_dict


def apply_styles(ws, style_dict: dict, style_ranges: dict, autofit: bool = True):
    # Register named styles to the workbook
    for named_style in style_dict.values():
        if named_style.name not in ws.parent.named_styles:
            ws.parent.add_named_style(named_style)

    console.print("Applying named styles...")
    # Apply other styles
    for style_name, ranges in style_ranges.items():
        if style_name in style_dict:  # Check if the style is defined
            for cell_range in ranges:
                range_split = cell_range.split(":")
                min_cell = coordinate_from_string(range_split[0])
                min_col, min_row = column_index_from_string(min_cell[0]), min_cell[1]
                if len(range_split) > 1:  # If the range includes more than one cell
                    max_cell = coordinate_from_string(range_split[1])
                    max_col, max_row = (
                        column_index_from_string(max_cell[0]),
                        max_cell[1],
                    )
                else:  # If the range includes only one cell
                    max_col, max_row = min_col, min_row
                for row in ws.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    for cell in row:
                        cell.style = style_dict[style_name]

    # Zoom out
    ws.sheet_view.zoomScale = 75

    # Autoadjust width for each column depending on its content
    if autofit:
        auto_format_cell_width(ws)


def create_ddb_table(
    df: DataFrame,
    db_file: str,
    table_name: str,
    table_schema: str,
    query_file: str,
    insert_instead: bool = False,
):
    console.print(f"Creating table into DuckDB file {db_file}...")

    with duckdb.connect(db_file) as db:
        temp_view_register = f"{table_name}_temp"
        temp_table_name = "tmpoffers_deleteafter"
        db.register(temp_view_register, df)
        if not all([table_name, query_file]):
            console.print("table_name and query_file must be specified before running.")
            return
        db.execute(
            f"create or replace table {temp_table_name} as select * from {temp_view_register}"
        )

        # Read file and split queries
        console.print("Fixing data...")
        with open(query_file, "r", encoding="utf8") as f:
            queries = f.read().split(";")

        # Iterate over each query
        for query in queries:
            # Skip empty queries
            if not query.strip():
                continue
            if insert_instead:
                # Replace placeholders with parameters
                query = query.replace("{table_schema}.", "").replace(
                    "{table_name}", temp_table_name
                )
            else:
                query = query.replace("{table_schema}.", "").replace(
                    "{table_name}", temp_table_name
                )
            # Execute query
            console.print("Executing query:", query)
            db.execute(query)

        if insert_instead:
            console.print(
                f"Inserting values into table {table_name} in {table_schema} from temp data..."
            )
            db.execute(
                f"insert into {table_schema}.{table_name} by name (select * from {temp_table_name})"
            )
            return

        console.print(
            f"Creating table {table_name} in {table_schema} from temp data..."
        )
        db.execute(
            f"create or replace table {table_schema}.{table_name} as select * from {temp_table_name}"
        )
        db.execute(f"drop table {temp_table_name}")

    return


def write_offers(
    output_file: Path,
    config: FileSettings,
    style_specs: dict[str, None],
    **kwargs,
):
    """
    Writes the data to an output Excel workbook.
    """
    with duckdb.connect(config.db_file) as db:
        with open("./queries/write_offers.sql", "r", encoding="utf8") as f:
            query = f.read()
            # Execute query
            dataframe = db.sql(query).pl()
    no_of_files, no_of_variables = dataframe.shape

    columns_start_at = 1
    first_column_letter, last_column_letter = (
        get_column_letter(columns_start_at),
        get_column_letter(no_of_variables),
    )
    # Define your custom formatting schema here
    entire_range = f"{first_column_letter}{config.header_start}:{last_column_letter}{config.header_start + no_of_files}"
    header_range = f"{first_column_letter}{config.header_start}:{last_column_letter}{config.header_start}"

    cell_ranges = {
        "default": [entire_range],
        "header": [header_range],
        "data": [
            f"E{config.header_start + 1}:E{config.header_start + no_of_files}",
            f"H{config.header_start + 1}:H{config.header_start + no_of_files}",
            f"K{config.header_start + 1}:M{config.header_start + no_of_files}",
            f"S{config.header_start + 1}:V{config.header_start + no_of_files}",
        ],
        "dates": [
            f"F{config.header_start + 1}:F{config.header_start + no_of_files}",
            f"I{config.header_start + 1}:I{config.header_start + no_of_files}",
            f"R{config.header_start + 1}:R{config.header_start + no_of_files}",
        ],
        "input": ["B3"],
        "title": ["A1"],
        "subtitle": ["A3"],
    }

    config.areas_to_style = cell_ranges

    workbook = openpyxl.Workbook()
    # Remove the default sheet created and add new sheets as per data keys
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("Offers Data")

    total_rows, total_columns = dataframe.shape
    last_column_as_letter = get_column_letter(total_columns)

    # Writing workbook creation date, title and description
    title = config.sheet_name
    start_row = config.header_start

    sheet.cell(column=1, row=1, value=title)
    sheet.cell(column=1, row=2, value="Created on:")
    sheet.cell(column=2, row=2, value=datetime.datetime.now())
    sheet.cell(column=1, row=3, value="Created by:")
    sheet.cell(column=2, row=3, value=os.environ.get("USERNAME"))
    sheet.cell(
        column=1,
        row=(start_row - 1),
        value=f"=COUNTA(A{start_row + 1}:A{start_row + total_rows})",
    )

    # Writing data from dataframe to sheet starting from start_row
    for i, row in enumerate(
        dataframe_to_rows(dataframe.to_pandas(), index=False, header=True), 1
    ):
        for j, cell in enumerate(row, 1):
            cell = str(tuple(cell)) if isinstance(cell, list) else cell
            sheet.cell(row=i + start_row - 1, column=j, value=cell)

    autofit_check = kwargs.get("autofit", True)
    apply_styles(sheet, style_specs, config.areas_to_style, autofit_check)
    filters = sheet.auto_filter
    filters.ref = f"A{start_row}:{last_column_as_letter}{total_rows}"
    sheet.freeze_panes = f"B{start_row + 1}"

    console.print("Saving output file")
    workbook.save(output_file)
    console.print(f"File saved in: {output_file}")


def timing(f):
    def wrap(*args, **kwargs):
        start_time = time.perf_counter()
        ret = f(*args, **kwargs)
        end_time = time.perf_counter() - start_time
        console.print(f"Total time elapsed: {end_time:0.3f} seconds")
        return ret

    return wrap
