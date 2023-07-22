from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from rich.console import Console
import re
import pandas as pd
import duckdb
import argparse
from conf.functions import (create_style, apply_styles)
from conf.settings import (FileSettings, offersconf, stockconf, pipeconf, QUERIES_DIR)

# WARN: Deberíamos admitir una clase FileSettings como input, y de ahi sacar los parametros
# Como strat_sheet, output_file, etc.

configuration = None
styles_first_row = 3
styles_first_column = 2

console = Console()

def insert_dataframes_to_workbook(dataframes_dict: dict, settings: FileSettings) -> None:
    workbook = load_workbook(settings.get_output_path())
    # Remove the default sheet created and add new sheets as per data keys
    try:
        workbook.remove(workbook[settings.strat_sheet])
        workbook.create_sheet(settings.strat_sheet)
    except KeyError:
        workbook.create_sheet(settings.strat_sheet)
    sheet = workbook[settings.strat_sheet]
    for title, (df, start_row, start_column) in dataframes_dict.items():
        console.print(f"Adding {title} to file...")
        rows = dataframe_to_rows(df, index=False, header=True)
        sheet.cell(row=start_row - 1, column=start_column, value=title)
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, value in enumerate(row, start=start_column):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    # style_spec = create_style(settings.styles_file)
    # apply_styles(sheet, style_spec, settings.areas_to_style)
    console.print(f"Saving file: {settings.get_output_path()}")
    workbook.save(settings.get_output_path())


def parse_sql_file(file_path: str) -> dict:
    queries = {}
    with open(file_path, "r") as file:
        content = file.read()
        queries_list = re.split(r"--\s*(\w+)\s*\n", content)[1:]
        queries = {queries_list[i]: queries_list[i+1] for i in range(0, len(queries_list), 2)}
    return queries


def get_strat_from_query(query: str, connection: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    return connection.execute(query).df()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--file",
        choices=["offers", "pipeline", "stock"],
        required=True,
        help="Whether or not to scan the update directory and update the offer data. Setting this to TRUE without the current_year option will update latest offers (2023 in this case)"
    )
    args = parser.parse_args()
    match args.file:
        case "offers":
            configuration = offersconf
            query_file = Path(QUERIES_DIR) / "strats_offers.sql"
        case "pipeline":
            configuration = pipeconf
            query_file = Path(QUERIES_DIR) / "strats_pipe.sql"
        case "stock":
            configuration = stockconf
            query_file = Path(QUERIES_DIR) / "strats_stock.sql"

    first_dict = parse_sql_file(query_file.as_posix())
    output_dict = {}

    with duckdb.connect(configuration.db_file) as db:
        current_row = styles_first_row
        for k, v in first_dict.items():
            df = get_strat_from_query(v, db)
            vertical_size = df.shape[0] 
            output_dict[k] = (get_strat_from_query(v, db), current_row, styles_first_column)
            current_row = current_row + vertical_size + 4
    insert_dataframes_to_workbook(output_dict, configuration)
