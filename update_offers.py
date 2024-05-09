import argparse
import datetime
import os
import re
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd
import polars as pl
import pendulum as pdl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from rich.console import Console

from conf.functions import (
    apply_styles,
    create_ddb_table,
    create_style,
    find_files_included,
    load_json_config,
    write_offers,
)
from conf.settings import cell_address_file
from conf.settings import offersconf as conf
from conf.settings import sap_mapping_file, styles_file

# Instanciar la consola bonita
console = Console()

# Instanciar las variables de ficheros, carpetas y otros
stylesheet = create_style(styles_file)


def extract_cell_values(file: str, search_strings: dict, columns_dict: dict):
    """
    Opens an Excel file and converts the worksheet to a dictionary. It then looks for
    specific strings, moves to the relative cell addresses based on both x and y offsets,
    and extracts the value at that location. Also extracts unique values from specified
    columns in a table on another sheet.
    Includes the file name in the returned data.
    """
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        console.print(f"Error when loading file {file}. Details: {e}")
        data = {}
        data["read_status"] = "Fail"
        data["read_details"] = str(e)
        data["full_path"] = Path(file).as_posix()
        data["file_name"] = os.path.basename(file)
        return data

    data = {label: None for label in search_strings.keys()}
    data["full_path"] = os.path.abspath(file)
    data["file_name"] = os.path.basename(file)

    try:
        hoja_ficha = list(
            filter(
                lambda z: re.match("ficha", z, flags=re.IGNORECASE), workbook.sheetnames
            )
        )
        hoja_ficha = hoja_ficha.pop() if hoja_ficha else "FICHA"
        sheet = workbook[hoja_ficha]
        data["read_status"] = "Success"
        data["read_details"] = None
    except KeyError as e:
        data["read_status"] = "Fail"
        data["read_details"] = e
        return data

    # Convert the worksheet to a dictionary for faster searching
    sheet_dict = {
        (cell.row, cell.column): cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }

    for (cell_row, cell_col), cell_value in sheet_dict.items():
        for label, [regex, offset_y, offset_x] in search_strings.items():
            if re.match(
                regex, str(cell_value), re.IGNORECASE
            ):  # compare values using the regular expression
                try:
                    target_cell_address = (
                        cell_row - offset_y,
                        cell_col + offset_x,
                    )
                    target_cell_value = sheet_dict.get(target_cell_address)
                    if target_cell_value is not None:
                        data[label] = target_cell_value
                        break
                except IndexError:
                    continue
                else:
                    break
            else:
                continue

    # Load the JSON file for table columns and process table columns if specified
    if columns_dict:
        regex = re.compile(r"\b(?:SAP|^Oferta)\b", re.IGNORECASE)
        posibles_hojas = list(filter(regex.search, workbook.sheetnames))
        hojas_sin_imagenes = []
        for sh in posibles_hojas:
            if not workbook[sh]._images:
                hojas_sin_imagenes.append(sh)
        try:
            hoja_detalle = hojas_sin_imagenes[0]
        except IndexError:
            hoja_detalle = posibles_hojas[0]
        try:
            table_sheet = workbook[hoja_detalle]
            for col in columns_dict:
                # Assume the first row contains headers
                header_row = next(
                    table_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )
                # Get index of the target column based on header
                try:
                    col_idx = header_row.index(col)
                except ValueError:
                    # console.print(
                    #     f'Warning: Column header "{col}" not found in sheet. Skipping this column.'
                    # )
                    continue
                # Extract unique values from the column
                column_values = [
                    cell[col_idx]
                    if not isinstance(cell[col_idx], str) or not cell[col_idx].isdigit()
                    else int(cell[col_idx].lstrip("0"))
                    for cell in table_sheet.iter_rows(min_row=2, values_only=True)
                ]
                # Remove duplicates by converting to a set, after filtering out the null values
                unique_values = set(filter(None, column_values))
                # If there's only one unique value, unpack it from the set
                if len(unique_values) == 1:
                    unique_values = unique_values.pop()
                    if isinstance(unique_values, str) and unique_values.isdigit():
                        unique_values = int(
                            unique_values.lstrip("0")
                        )  # Remove leading zeros for numeric strings
                else:
                    # Convert to list and sort to make output more predictable
                    try:
                        unique_values = sorted(unique_values)
                    except TypeError as e:
                        console.print(f"Error en fichero {file} debido a {e}")
                        data["read_status"] = "Warning"
                        data["read_details"] = str(e)
                        continue
                # Add the unique values to the output
                data[columns_dict.get(col)] = unique_values if unique_values else None
        except KeyError as e:
            console.print(f"{e}\nFichero causante: {file}")
            data["read_status"] = "Warning"
            data["read_details"] = str(e)

        # Cleanup of bad data or strings
        if data["offer_date"] is not None and not isinstance(
            data["offer_date"], datetime.datetime
        ):
            try:
                data["offer_date"] = pdl.parse(
                    data["offer_date"], strict=False
                ).to_date_string()
            except Exception as e:
                console.print(
                    f"Error al identificar la fecha de la oferta: >> {e}. Fichero: {file}"
                )
                data["offer_date"] = None

    return data


def load_previous_data(
    sheet: str = conf.sheet_name, start_row: int = conf.header_start
) -> pd.DataFrame | bool:
    latest_file_name_pattern = conf.output_file
    matching_files = []
    for f in Path(conf.output_dir).glob("*.*"):
        if re.search(latest_file_name_pattern, f.name):
            matching_files.append(f)

    sorted_files = sorted([f for f in matching_files], key=os.path.getmtime)
    previous_file = sorted_files[-1]
    if os.path.isfile(previous_file):
        return pd.read_excel(
            previous_file, sheet_name=sheet, skiprows=start_row - 1
        )  # load the existing data
    else:
        # return False for the case where no previous data exists
        return False


def enrich_offers(input_query: str, reuse_latest_file: bool = False):
    console.print("Getting data from portfolio management...")
    with open("./queries/unnest_unique_urs.sql", encoding="utf8") as sql_file:
        query = sql_file.read()
        improved_query = f"CREATE TEMP TABLE unnested_data AS {query}"

    with duckdb.connect(conf.db_file) as db:
        # Ingest previous data
        db.execute("""set global pandas_analyze_sample=10000""")
        db.execute(improved_query)

        if reuse_latest_file:
            console.print(
                "Opening latest offers file and retrieving additional columns..."
            )
            previous_data = load_previous_data(conf.sheet_name)
            if previous_data:
                db.register("tmp_enrich_df", previous_data)
                excluded_columns_from_origin = [
                    "unique_urs",
                    "commercialdevs",
                    "jointdevs",
                    "offer_id",
                ]
            else:
                db.execute(f"""create temp table tmp_enrich_df as {input_query}""")
                excluded_columns_from_origin = [
                    "unique_urs",
                    "commercialdev",
                    "jointdev",
                    "offer_id",
                ]
        else:
            db.execute(f"""create temp table tmp_enrich_df as {input_query}""")
            excluded_columns_from_origin = [
                "unique_urs",
                "commercialdev",
                "jointdev",
                "offer_id",
            ]

        # Create enriched table, for later usage
        db.execute(
            f"""
            create or replace table offers_enriched_table as (
            with all_data as (
            select  t.* exclude({",".join(excluded_columns_from_origin) if len(excluded_columns_from_origin) > 1 else excluded_columns_from_origin.pop()}),
                    u.* exclude(unique_id)
            from tmp_enrich_df t
            left join unnested_data u
            on t.unique_id = u.unique_id
            ),
            filtered_columns as (
            select columns(x -> x not similar to '.+:1')
            from all_data)
            select * from filtered_columns);
            """
        )

        expanded_df = db.execute(
            """select unique_id, offer_id, * exclude(unique_id, offer_id) from offers_enriched_table order by offer_date desc"""
        ).df()

    return expanded_df


def write_output(
    output_file: str,
    data: dict[str, pl.DataFrame],
    style_specs: str,
    style_ranges: str,
    start_row: int,
    title: str,
    **kwargs,
):
    """
    Writes the data to an output Excel workbook.
    """
    workbook = openpyxl.Workbook()
    # Remove the default sheet created and add new sheets as per data keys
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name, dataframe in data.items():
        sheet = workbook.create_sheet(sheet_name)
        total_rows, total_columns = dataframe.shape
        last_column_as_letter = get_column_letter(total_columns)

        # Writing workbook creation date, title and description
        sheet.cell(column=1, row=1, value=title)
        sheet.cell(column=1, row=2, value="Created on:")
        sheet.cell(column=2, row=2, value=datetime.datetime.now())
        sheet.cell(column=1, row=3, value="Created by:")
        sheet.cell(column=2, row=3, value=os.environ.get("USERNAME"))
        sheet.cell(
            column=1,
            row=(start_row - 1),
            value=f"=COUNTA(A{start_row + 1}:A{start_row + total_rows})",
        )

        # Writing data from dataframe to sheet starting from start_row
        for i, row in enumerate(
            dataframe_to_rows(dataframe.to_pandas(), index=False, header=True), 1
        ):
            for j, cell in enumerate(row, 1):
                cell = str(tuple(cell)) if isinstance(cell, list) else cell
                sheet.cell(row=i + start_row - 1, column=j, value=cell)

        autofit_check = kwargs.get("autofit", True)
        apply_styles(sheet, style_specs, style_ranges, autofit_check)
        filters = sheet.auto_filter
        filters.ref = f"A{start_row}:{last_column_as_letter}{total_rows}"
        sheet.freeze_panes = f"B{start_row + 1}"

    console.print("Saving output file")
    workbook.save(output_file)
    console.print(f"File saved in: {output_file}")


def main(
    update_offers: bool = False,
    write_file: bool = False,
    refresh_data: bool = False,
    year_to_scrape: int = datetime.datetime.now().year,
):
    if update_offers:
        # Create the output directory if not exists
        console.print(f"Creating path to files: {conf.output_dir}")
        Path(conf.output_dir).mkdir(exist_ok=True)

        table_name_to_use = "ws_offers_"
        ddb_table_name = table_name_to_use + str(year_to_scrape)

        # Extract the workbooks information one by one, then append the dictionary records to a 'data' variable
        cell_addresses = load_json_config(cell_address_file)
        sap_columns_mapping = load_json_config(sap_mapping_file)

        folder_pattern = rf"{year_to_scrape}"
        files, total_files = find_files_included(
            conf.directory, folder_pattern, conf.db_file, only_new_files=refresh_data
        )
        files_count = len(files)
        if files_count == 0:
            console.print("No files (new or old) found. Aborting...")
            return

        if files_count == total_files:
            console.print(
                "The source files are the same, rerun the script without --refresh."
            )
            return

        with console.status("Extracting cell values from files...") as status:
            data = []
            for idx, file in enumerate(files, start=1):
                status.update(
                    f"[{idx}/{files_count}] ~ Loading data from file: [bold green]{file}[/bold green]"
                )
                data.append(
                    extract_cell_values(file, cell_addresses, sap_columns_mapping)
                )
        # Use the 'data' variable to create a DataFrame structure which will be manipulated/modified
        console.print("Assembling offer data into a DataFrame...")
        df = pd.DataFrame(data)
        create_ddb_table(
            df,
            conf.db_file,
            query_file="./queries/fix_offers.sql",
            table_name=ddb_table_name,
            table_schema=conf.db_schema,
            insert_instead=refresh_data,
        )

    if write_file:
        write_offers(
            conf.get_output_path(),
            conf,
            stylesheet,
            # reuse_latest_file=True,
            autofit=False,
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--update",
        default=False,
        action="store_true",
        help="Whether or not to scan the update directory and update the offer data. Setting this to TRUE without the current_year option will update latest offers",
    )
    parser.add_argument(
        "--year",
        default=datetime.datetime.now().year,
        help="Year for the offers to scan",
    )
    parser.add_argument(
        "--fix",
        default=False,
        action="store_true",
        help="Use the latest offers file as base and bring any custom data columns that may have been added",
    )
    parser.add_argument(
        "--write",
        default=False,
        action="store_true",
        help="Create the Excel file",
    )
    parser.add_argument(
        "--refresh",
        default=False,
        action="store_true",
        help="Only insert new files",
    )
    args = parser.parse_args()
    main(
        update_offers=args.update,
        write_file=args.write,
        year_to_scrape=args.year,
        refresh_data=args.refresh,
    )
